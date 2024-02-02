import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_file_2022 = ('2022_年間売上表.xlsx', 'Sheet1')
input_file_2023 = ('2023_年間売上表.xlsx', 'Sheet1')
output_file = '売上集計表.xlsx'

# ①データの連結
df_2022 = pd.read_excel(input_file_2022[0], sheet_name=input_file_2022[1])
df_2023 = pd.read_excel(input_file_2023[0], sheet_name=input_file_2023[1])
df_combined = pd.concat([df_2022, df_2023], ignore_index=True)

# ②データの集計
df_aggregated = df_combined.groupby(['商品', '売上年']).agg({'金額（千円）': 'sum'}).reset_index()

# ③売上集計表の作成と書き込み
df_aggregated.to_excel(output_file, index=False)

# ④ヘッダーの書式変更
workbook = load_workbook(output_file)
worksheet = workbook.active

# ヘッダー部分を薄いグレーに設定
for cell in worksheet[1]:
    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# 幅調整
worksheet.column_dimensions['A'].width = 20
worksheet.column_dimensions['B'].width = 12
worksheet.column_dimensions['C'].width = 12

# 保存
workbook.save(output_file)
